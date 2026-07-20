from __future__ import annotations

from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
import csv
import json
import re
import uuid

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter, landscape
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    Flowable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .pdf_reports import REPORT_DIR


GREEN = "168345"
DARK_GREEN = colors.HexColor("#0E5D32")
LIGHT_GREEN = colors.HexColor("#EAF5EE")


TEMPLATE_DESCRIPTIONS = {
    "executive": (
        "High-level KPIs, totals, notable findings, and a limited supporting "
        "sample for leadership review."
    ),
    "detailed": (
        "Complete operational records with useful IDs, dates, statuses, "
        "amounts, and source fields."
    ),
    "customer_facing": (
        "Clean customer-ready presentation with internal IDs, API names, "
        "queries, and technical metadata removed."
    ),
    "audit": (
        "Full traceability including dataset IDs, source, original query, "
        "intent, timestamps, complete records, and raw payload evidence."
    ),
}


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip()).strip("-")
    return cleaned[:80] or "ribbit-report"


def _friendly_label(value: str) -> str:
    value = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", value)
    value = value.replace("_", " ")
    replacements = {
        "id": "ID",
        "api": "API",
        "url": "URL",
        "psa": "PSA",
    }
    words = []
    for word in value.split():
        words.append(replacements.get(word.casefold(), word.capitalize()))
    return " ".join(words)


def _is_internal_field(key: str) -> bool:
    lowered = key.casefold()
    return (
        lowered == "id"
        or lowered.endswith("_id")
        or lowered.endswith("id")
        or lowered
        in {
            "source",
            "intent",
            "conversation_id",
            "dataset_id",
            "dataset_type",
            "dataset_title",
            "raw",
            "soap_fallback_error",
            "ledger_source",
        }
    )


def _flatten_record(
    record: dict[str, Any],
    *,
    template: str,
) -> dict[str, Any]:
    flat: dict[str, Any] = {}

    for key, value in record.items():
        if key == "raw" and template != "audit":
            continue

        if template == "customer_facing" and _is_internal_field(key):
            continue

        if template == "executive" and (
            _is_internal_field(key)
            or key.casefold()
            in {
                "created_at",
                "modified_at",
                "start_date",
                "end_date",
                "raw",
            }
        ):
            continue

        output_key = (
            _friendly_label(key)
            if template == "customer_facing"
            else key
        )

        if isinstance(value, (dict, list)):
            flat[output_key] = json.dumps(
                value,
                ensure_ascii=False,
                default=str,
            )
        else:
            flat[output_key] = value

    return flat


def _dataset_records(
    dataset: dict[str, Any],
) -> list[dict[str, Any]]:
    data = dataset.get("data") or {}
    data_type = dataset.get("data_type")
    keys = {
        "ccwr_renewals": "ccwr_renewals",
        "billing_ledger": "ledger_entries",
        "service_products": "service_products",
        "tickets": "tickets",
        "projects": "projects",
        "invoices": "invoices",
        "opportunities": "opportunities",
        "contacts": "contacts",
        "customers": "customers",
        "project_activity": "activity",
        "customer": "customer",
    }
    value = data.get(keys.get(data_type, ""))

    if isinstance(value, list):
        return [
            item for item in value
            if isinstance(item, dict)
        ]
    if isinstance(value, dict):
        return [value]
    return []


def _number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(
            str(value)
            .replace("$", "")
            .replace(",", "")
        )
    except (TypeError, ValueError):
        return 0.0


def _first(record: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        value = record.get(key)
        if value not in (None, ""):
            return value
    return None


def _dataset_insights(
    dataset: dict[str, Any],
) -> list[tuple[str, str]]:
    data = dataset.get("data") or {}
    data_type = dataset.get("data_type")
    records = _dataset_records(dataset)
    insights: list[tuple[str, str]] = [
        ("Records", f"{len(records):,}"),
    ]

    if data_type == "ccwr_renewals":
        summary = data.get("renewal_summary") or {}
        return [
            (
                "Subscriptions",
                f"{int(summary.get('total_subscriptions') or 0):,}",
                "Cisco subscriptions returned",
            ),
            (
                "Active",
                f"{int(summary.get('active') or 0):,}",
                "Active subscriptions",
            ),
            (
                "Actionable overdue",
                f"{int(summary.get('actionable_overdue') or 0):,}",
                "Active renewal date has passed",
            ),
            (
                "Due in 30 days",
                f"{int(summary.get('due_0_30') or 0):,}",
                "Immediate renewals",
            ),
            (
                "Due in 31-60",
                f"{int(summary.get('due_31_60') or 0):,}",
                "Near-term renewals",
            ),
            (
                "Due in 61-90",
                f"{int(summary.get('due_61_90') or 0):,}",
                "Upcoming renewals",
            ),
            (
                "Closed",
                f"{int(summary.get('closed') or 0):,}",
                "Expired or cancelled",
            ),
            (
                "High risk",
                f"{int(summary.get('high_risk') or 0):,}",
                "Due within 30 days",
            ),
        ]

    if data_type == "billing_ledger":
        summary = data.get("ledger_summary") or {}
        insights.extend(
            [
                (
                    "Total charges",
                    f"${_number(summary.get('total_charges')):,.2f}",
                ),
                (
                    "Total credits",
                    f"${_number(summary.get('total_credits')):,.2f}",
                ),
                (
                    "Charges less credits",
                    f"${_number(summary.get('net_charges_less_credits')):,.2f}",
                ),
            ]
        )
    elif data_type == "tickets":
        statuses = Counter(
            str(_first(row, ["status", "Status"]) or "Unknown")
            for row in records
        )
        oldest = max(
            (
                _number(
                    _first(
                        row,
                        ["age_days", "ageDays", "AgeDays"],
                    )
                )
                for row in records
            ),
            default=0,
        )
        insights.extend(
            [
                ("Statuses", f"{len(statuses):,}"),
                ("Oldest ticket", f"{oldest:,.0f} days"),
                (
                    "Most common status",
                    statuses.most_common(1)[0][0]
                    if statuses
                    else "Not available",
                ),
            ]
        )
    elif data_type == "invoices":
        total = sum(
            _number(
                _first(
                    row,
                    ["amount", "Amount", "total", "Total"],
                )
            )
            for row in records
        )
        balance = sum(
            _number(
                _first(
                    row,
                    ["balance", "Balance", "amountDue"],
                )
            )
            for row in records
        )
        insights.extend(
            [
                ("Invoice total", f"${total:,.2f}"),
                ("Open balance", f"${balance:,.2f}"),
            ]
        )
    elif data_type == "opportunities":
        total = sum(
            _number(_first(row, ["amount", "Amount"]))
            for row in records
        )
        insights.append(
            ("Pipeline value", f"${total:,.2f}")
        )
    elif data_type == "projects":
        statuses = Counter(
            str(
                _first(
                    row,
                    [
                        "projectStatusName",
                        "statusName",
                        "status",
                    ],
                )
                or "Unknown"
            )
            for row in records
        )
        insights.extend(
            [
                ("Project statuses", f"{len(statuses):,}"),
                (
                    "Most common status",
                    statuses.most_common(1)[0][0]
                    if statuses
                    else "Not available",
                ),
            ]
        )
    elif data_type == "service_products":
        active = sum(
            1
            for row in records
            if str(
                _first(row, ["status", "Status"]) or ""
            ).casefold()
            == "active"
        )
        insights.append(("Active services", f"{active:,}"))

    return insights


def _summary_lines(
    datasets: list[dict[str, Any]],
    *,
    template: str,
) -> list[str]:
    total_records = sum(
        int(item.get("record_count") or 0)
        for item in datasets
    )
    lines = [
        f"Datasets included: {len(datasets)}",
        f"Total records included: {total_records:,}",
    ]

    for dataset in datasets:
        insights = ", ".join(
            f"{label}: {value}"
            for label, value in _dataset_insights(dataset)
        )
        lines.append(
            f"{dataset.get('title', 'Dataset')} - {insights}"
        )

    if template == "audit":
        lines.append(
            "Audit mode includes source metadata, original query, intent, "
            "dataset identifiers, and raw evidence."
        )
    elif template == "customer_facing":
        lines.append(
            "Customer-facing mode excludes internal identifiers and "
            "technical source metadata."
        )
    elif template == "executive":
        lines.append(
            "Executive mode emphasizes decisions and KPIs rather than "
            "record-level operational detail."
        )

    return lines


def _record_limit(
    template: str,
    *,
    include_raw_records: bool,
) -> int | None:
    if template == "executive":
        return 10
    if template == "customer_facing":
        return 100 if include_raw_records else 25
    if template == "detailed":
        return None if include_raw_records else 100
    return None


def create_report(
    *,
    datasets: list[dict[str, Any]],
    title: str,
    report_format: str,
    template: str,
    include_summary: bool,
    include_raw_records: bool,
) -> dict[str, str]:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report_id = str(uuid.uuid4())
    stem = f"{_safe_filename(title)}-{report_id[:8]}"

    if report_format == "xlsx":
        path = REPORT_DIR / f"{stem}.xlsx"
        _create_xlsx(
            path,
            datasets=datasets,
            title=title,
            template=template,
            include_summary=include_summary,
            include_raw_records=include_raw_records,
        )
    elif report_format == "csv":
        path = REPORT_DIR / f"{stem}.csv"
        _create_csv(
            path,
            datasets=datasets,
            template=template,
            include_raw_records=include_raw_records,
        )
    else:
        path = REPORT_DIR / f"{stem}.pdf"
        _create_pdf(
            path,
            datasets=datasets,
            title=title,
            template=template,
            include_summary=include_summary,
            include_raw_records=include_raw_records,
        )

    return {
        "report_id": report_id,
        "download_url": f"/api/downloads/{path.name}",
        "download_name": path.name,
        "format": report_format,
    }


def _style_summary_sheet(
    ws,
    *,
    title: str,
    template: str,
    datasets: list[dict[str, Any]],
) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=GREEN,
    )
    ws["A2"] = (
        f"Generated "
        f"{datetime.now(timezone.utc).strftime('%B %d, %Y %I:%M %p UTC')}"
    )
    ws["A4"] = "Report type"
    ws["B4"] = template.replace("_", " ").title()
    ws["A5"] = "Purpose"
    ws["B5"] = TEMPLATE_DESCRIPTIONS[template]
    ws["B5"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 95

    row = 7
    for line in _summary_lines(
        datasets,
        template=template,
    ):
        ws.cell(row=row, column=1, value="•")
        ws.cell(row=row, column=2, value=line)
        ws.cell(row=row, column=2).alignment = Alignment(
            wrap_text=True
        )
        row += 1


def _create_xlsx(
    path: Path,
    *,
    datasets: list[dict[str, Any]],
    title: str,
    template: str,
    include_summary: bool,
    include_raw_records: bool,
) -> None:
    wb = Workbook()
    ws = wb.active

    ccwr_only = (
        bool(datasets)
        and all(
            dataset.get("data_type") == "ccwr_renewals"
            for dataset in datasets
        )
        and template in {"executive", "customer_facing"}
    )
    if ccwr_only:
        _ccwr_xlsx_dashboard(
            wb,
            datasets=datasets,
            title=title,
        )
        wb.save(path)
        return

    ws.title = "Executive Summary" if template == "executive" else "Summary"
    _style_summary_sheet(
        ws,
        title=title,
        template=template,
        datasets=datasets,
    )

    if template == "audit":
        audit = wb.create_sheet("Audit Trail")
        headers = [
            "Dataset ID",
            "Conversation ID",
            "Title",
            "Data Type",
            "Source",
            "Original Query",
            "Intent",
            "Created At",
            "Record Count",
        ]
        for col, header in enumerate(headers, start=1):
            cell = audit.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=GREEN)

        for row_index, dataset in enumerate(datasets, start=2):
            values = [
                dataset.get("dataset_id"),
                dataset.get("conversation_id"),
                dataset.get("title"),
                dataset.get("data_type"),
                dataset.get("source"),
                dataset.get("query"),
                dataset.get("intent"),
                dataset.get("created_at"),
                dataset.get("record_count"),
            ]
            for col, value in enumerate(values, start=1):
                audit.cell(row=row_index, column=col, value=value)

        for col in range(1, len(headers) + 1):
            audit.column_dimensions[get_column_letter(col)].width = 24
        audit.column_dimensions["F"].width = 70
        audit.freeze_panes = "A2"

    for index, dataset in enumerate(datasets, start=1):
        records = _dataset_records(dataset)
        if not records:
            continue

        sheet_name = re.sub(
            r"[:\\/?*\[\]]",
            "",
            str(dataset.get("title") or f"Dataset {index}"),
        )
        sheet_name = sheet_name[:31] or f"Dataset {index}"
        if sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name[:27]} {index}"

        sheet = wb.create_sheet(sheet_name)
        limit = _record_limit(
            template,
            include_raw_records=include_raw_records,
        )
        selected_records = (
            records
            if limit is None
            else records[:limit]
        )
        rows = [
            _flatten_record(
                record,
                template=template,
            )
            for record in selected_records
        ]
        headers = list(
            dict.fromkeys(
                key
                for record in rows
                for key in record.keys()
            )
        )

        if not headers:
            sheet["A1"] = "No customer-safe fields were available."
            continue

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(
                row=1,
                column=col,
                value=header,
            )
            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )
            cell.fill = PatternFill(
                "solid",
                fgColor=GREEN,
            )
            cell.alignment = Alignment(
                wrap_text=True,
            )

        for row_index, record in enumerate(rows, start=2):
            for col, header in enumerate(headers, start=1):
                sheet.cell(
                    row=row_index,
                    column=col,
                    value=record.get(header),
                )

        for col, header in enumerate(headers, start=1):
            width = max(
                len(str(header)),
                max(
                    (
                        len(str(record.get(header, "")))
                        for record in rows[:200]
                    ),
                    default=0,
                ),
            )
            sheet.column_dimensions[
                get_column_letter(col)
            ].width = min(
                max(width + 2, 12),
                45,
            )

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        if template == "audit":
            raw_sheet_name = f"Raw {index}"[:31]
            raw_sheet = wb.create_sheet(raw_sheet_name)
            raw_sheet["A1"] = "Dataset ID"
            raw_sheet["B1"] = dataset.get("dataset_id")
            raw_sheet["A2"] = "Raw JSON"
            raw_sheet["B2"] = json.dumps(
                dataset.get("data") or {},
                ensure_ascii=False,
                default=str,
            )
            raw_sheet["B2"].alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )
            raw_sheet.column_dimensions["A"].width = 18
            raw_sheet.column_dimensions["B"].width = 120

    wb.save(path)


def _create_csv(
    path: Path,
    *,
    datasets: list[dict[str, Any]],
    template: str,
    include_raw_records: bool,
) -> None:
    rows: list[dict[str, Any]] = []
    limit = _record_limit(
        template,
        include_raw_records=include_raw_records,
    )

    for dataset in datasets:
        records = _dataset_records(dataset)
        selected_records = (
            records
            if limit is None
            else records[:limit]
        )

        for record in selected_records:
            flat = _flatten_record(
                record,
                template=template,
            )

            if template == "audit":
                flat = {
                    "dataset_id": dataset.get("dataset_id"),
                    "conversation_id": dataset.get("conversation_id"),
                    "dataset_title": dataset.get("title"),
                    "data_type": dataset.get("data_type"),
                    "source": dataset.get("source"),
                    "original_query": dataset.get("query"),
                    "intent": dataset.get("intent"),
                    "dataset_created_at": dataset.get("created_at"),
                    **flat,
                    "raw_dataset_json": json.dumps(
                        dataset.get("data") or {},
                        ensure_ascii=False,
                        default=str,
                    ),
                }
            elif template != "customer_facing":
                flat = {
                    "dataset_title": dataset.get("title"),
                    "data_type": dataset.get("data_type"),
                    **flat,
                }

            rows.append(flat)

    headers = list(
        dict.fromkeys(
            key
            for row in rows
            for key in row.keys()
        )
    )
    with path.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=headers,
        )
        writer.writeheader()
        writer.writerows(rows)


def _pdf_styles():
    styles = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "RibbitTitle",
            parent=styles["Title"],
            textColor=DARK_GREEN,
            fontSize=22,
            leading=26,
            alignment=TA_CENTER,
        ),
        "heading": ParagraphStyle(
            "RibbitHeading",
            parent=styles["Heading2"],
            textColor=DARK_GREEN,
            fontSize=14,
            leading=18,
        ),
        "body": ParagraphStyle(
            "RibbitBody",
            parent=styles["BodyText"],
            fontSize=9,
            leading=13,
        ),
        "small": ParagraphStyle(
            "RibbitSmall",
            parent=styles["BodyText"],
            fontSize=7,
            leading=9,
        ),
    }


def _pdf_kpi_table(
    dataset: dict[str, Any],
    *,
    body_style,
) -> Table:
    insights = _dataset_insights(dataset)
    data = [
        [
            Paragraph(label, body_style),
            Paragraph(value, body_style),
        ]
        for label, value in insights
    ]
    table = Table(
        data,
        colWidths=[2.2 * inch, 2.2 * inch],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), LIGHT_GREEN),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#17351F")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BBD2C2")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return table



def _date_value(value: Any) -> datetime | None:
    if value in (None, ""):
        return None

    text = str(value).strip().replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        pass

    for pattern in (
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(text[:19], pattern)
        except ValueError:
            continue

    return None


def _status_counter(
    dataset: dict[str, Any],
) -> Counter[str]:
    records = _dataset_records(dataset)
    data_type = dataset.get("data_type")

    status_keys = {
        "ccwr_renewals": ["risk_level", "status"],
        "tickets": ["status", "Status"],
        "projects": [
            "projectStatusName",
            "statusName",
            "status",
            "state",
        ],
        "invoices": ["status", "Status"],
        "opportunities": [
            "stageName",
            "status",
            "opportunityStatus",
        ],
        "service_products": ["status", "Status"],
        "billing_ledger": ["entry_type", "Type"],
    }

    keys = status_keys.get(
        str(data_type),
        ["status", "Status", "state"],
    )

    return Counter(
        str(_first(record, keys) or "Unknown")
        for record in records
    )


def _monthly_series(
    dataset: dict[str, Any],
) -> tuple[list[str], list[float], list[float]]:
    """
    Return labels, primary values, secondary values for dashboard charts.

    Billing ledger: charges vs credits.
    Other datasets: record count vs value/amount when available.
    """
    records = _dataset_records(dataset)
    data_type = dataset.get("data_type")
    buckets: dict[str, list[float]] = {}

    date_keys = [
        "created_date",
        "createdDate",
        "CreatedDate",
        "created_at",
        "createdAt",
        "invoiceDate",
        "InvoiceDate",
        "date",
        "Date",
    ]

    for record in records:
        date = _date_value(_first(record, date_keys))
        if not date:
            continue

        month = date.strftime("%b %Y")
        values = buckets.setdefault(month, [0.0, 0.0])

        if data_type == "billing_ledger":
            amount = _number(
                _first(record, ["amount", "Amount"])
            )
            entry_type = str(
                _first(record, ["entry_type", "Type"])
                or "CHARGE"
            ).upper()
            if entry_type == "CREDIT":
                values[1] += abs(amount)
            else:
                values[0] += abs(amount)
        else:
            values[0] += 1
            values[1] += _number(
                _first(
                    record,
                    [
                        "amount",
                        "Amount",
                        "balance",
                        "Balance",
                        "value",
                        "Value",
                    ],
                )
            )

    if not buckets:
        return ["No dated records"], [0.0], [0.0]

    parsed = []
    for label, values in buckets.items():
        try:
            parsed_date = datetime.strptime(label, "%b %Y")
        except ValueError:
            parsed_date = datetime.min
        parsed.append((parsed_date, label, values))

    parsed.sort(key=lambda item: item[0])
    parsed = parsed[-12:]

    return (
        [item[1] for item in parsed],
        [item[2][0] for item in parsed],
        [item[2][1] for item in parsed],
    )


def _dashboard_metrics(
    dataset: dict[str, Any],
) -> list[tuple[str, str, str]]:
    """
    label, value, helper text
    """
    data = dataset.get("data") or {}
    records = _dataset_records(dataset)
    data_type = dataset.get("data_type")
    statuses = _status_counter(dataset)

    if data_type == "billing_ledger":
        summary = data.get("ledger_summary") or {}
        return [
            (
                "Ledger entries",
                f"{len(records):,}",
                "Charge and credit records",
            ),
            (
                "Total charges",
                f"${_number(summary.get('total_charges')):,.2f}",
                "Gross billed activity",
            ),
            (
                "Total credits",
                f"${_number(summary.get('total_credits')):,.2f}",
                "Credits issued",
            ),
            (
                "Net activity",
                f"${_number(summary.get('net_charges_less_credits')):,.2f}",
                "Charges less credits",
            ),
            (
                "Charge records",
                f"{statuses.get('CHARGE', 0):,}",
                "Individual charge lines",
            ),
            (
                "Credit records",
                f"{statuses.get('CREDIT', 0):,}",
                "Individual credit lines",
            ),
            (
                "Statements",
                f"{len(set(str(_first(row, ['bill_id', 'StatementID']) or '') for row in records if _first(row, ['bill_id', 'StatementID']) not in (None, ''))):,}",
                "Unique statement IDs",
            ),
            (
                "Services",
                f"{len(set(str(_first(row, ['service_id', 'LineID']) or '') for row in records if _first(row, ['service_id', 'LineID']) not in (None, ''))):,}",
                "Unique service IDs",
            ),
        ]

    if data_type == "tickets":
        ages = [
            _number(
                _first(
                    record,
                    ["age_days", "ageDays", "AgeDays"],
                )
            )
            for record in records
        ]
        return [
            ("Total tickets", f"{len(records):,}", "Records returned"),
            ("New", f"{statuses.get('New', 0):,}", "New tickets"),
            ("Open", f"{statuses.get('Open', 0):,}", "Open tickets"),
            (
                "On hold",
                f"{statuses.get('On-Hold', statuses.get('On Hold', 0)):,}",
                "Waiting tickets",
            ),
            (
                "Needs review",
                f"{statuses.get('Needs Reviewed', 0):,}",
                "Review required",
            ),
            (
                "Average age",
                f"{(sum(ages) / len(ages) if ages else 0):,.1f}d",
                "Average ticket age",
            ),
            (
                "Oldest ticket",
                f"{(max(ages) if ages else 0):,.0f}d",
                "Oldest open record",
            ),
            (
                "Engineers",
                f"{len(set(str(_first(row, ['assigned_engineer', 'assignedEngineer']) or 'Unassigned') for row in records)):,}",
                "Assigned resources",
            ),
        ]

    if data_type == "invoices":
        total = sum(
            _number(_first(row, ["amount", "Amount", "total"]))
            for row in records
        )
        balance = sum(
            _number(_first(row, ["balance", "Balance", "amountDue"]))
            for row in records
        )
        return [
            ("Invoices", f"{len(records):,}", "Invoices returned"),
            ("Invoice total", f"${total:,.2f}", "Total invoice value"),
            ("Open balance", f"${balance:,.2f}", "Remaining balance"),
            (
                "Paid",
                f"{statuses.get('Paid', statuses.get('PAID', 0)):,}",
                "Paid invoices",
            ),
            (
                "Open",
                f"{statuses.get('Open', statuses.get('OPEN', 0)):,}",
                "Open invoices",
            ),
            (
                "Overdue",
                f"{statuses.get('Overdue', statuses.get('OVERDUE', 0)):,}",
                "Past-due invoices",
            ),
        ]

    if data_type == "projects":
        return [
            ("Projects", f"{len(records):,}", "Projects returned"),
            (
                "Active",
                f"{sum(count for name, count in statuses.items() if name.casefold() in {'active', 'open', 'in progress'}):,}",
                "Active work",
            ),
            (
                "Statuses",
                f"{len(statuses):,}",
                "Distinct project states",
            ),
            (
                "Customers",
                f"{len(set(str(_first(row, ['customerName', 'customer_name']) or '') for row in records if _first(row, ['customerName', 'customer_name']) not in (None, ''))):,}",
                "Customers represented",
            ),
        ]

    if data_type == "opportunities":
        total = sum(
            _number(_first(row, ["amount", "Amount"]))
            for row in records
        )
        return [
            ("Opportunities", f"{len(records):,}", "Pipeline records"),
            ("Pipeline value", f"${total:,.2f}", "Total opportunity value"),
            ("Stages", f"{len(statuses):,}", "Pipeline stages"),
            (
                "Won",
                f"{sum(count for name, count in statuses.items() if 'won' in name.casefold()):,}",
                "Won opportunities",
            ),
        ]

    if data_type == "service_products":
        active = sum(
            count
            for name, count in statuses.items()
            if name.casefold() == "active"
        )
        total_rate = sum(
            _number(_first(row, ["rate", "Rate", "price", "Price"]))
            * max(_number(_first(row, ["quantity", "Quantity"])), 1)
            for row in records
        )
        return [
            ("Services", f"{len(records):,}", "Service products"),
            ("Active", f"{active:,}", "Currently active"),
            (
                "Inactive",
                f"{max(len(records) - active, 0):,}",
                "Inactive services",
            ),
            (
                "Estimated monthly",
                f"${total_rate:,.2f}",
                "Rate multiplied by quantity",
            ),
        ]

    return [
        ("Records", f"{len(records):,}", "Records returned"),
        ("Statuses", f"{len(statuses):,}", "Distinct states"),
        (
            "Primary status",
            statuses.most_common(1)[0][0]
            if statuses
            else "Unknown",
            "Most common state",
        ),
        (
            "Primary count",
            f"{statuses.most_common(1)[0][1]:,}"
            if statuses
            else "0",
            "Records in primary state",
        ),
    ]


class DashboardKPIGrid(Flowable):
    def __init__(
        self,
        metrics: list[tuple[str, str, str]],
        width: float,
    ):
        super().__init__()
        self.metrics = metrics[:8]
        self.width = width
        self.height = 2.22 * inch

    def wrap(self, availWidth, availHeight):
        return self.width, self.height

    def draw(self):
        canvas = self.canv
        cols = 4
        rows = 2
        gap = 7
        card_width = (
            self.width - gap * (cols - 1)
        ) / cols
        card_height = (
            self.height - gap * (rows - 1)
        ) / rows

        fills = [
            colors.HexColor("#FFF0D2"),
            colors.HexColor("#F7F8FA"),
            colors.HexColor("#12B9D0"),
            colors.HexColor("#10AFC2"),
            colors.HexColor("#F7F8FA"),
            colors.HexColor("#F7F8FA"),
            colors.HexColor("#F7F8FA"),
            colors.HexColor("#F7F8FA"),
        ]

        for index, (label, value, helper) in enumerate(
            self.metrics
        ):
            row = index // cols
            col = index % cols
            x = col * (card_width + gap)
            y = self.height - (row + 1) * card_height - row * gap

            fill = fills[index % len(fills)]
            canvas.setFillColor(fill)
            canvas.setStrokeColor(colors.HexColor("#D6DBDF"))
            canvas.roundRect(
                x,
                y,
                card_width,
                card_height,
                5,
                fill=1,
                stroke=1,
            )

            dark_text = (
                fill != colors.HexColor("#12B9D0")
                and fill != colors.HexColor("#10AFC2")
            )
            text_color = (
                colors.HexColor("#20262D")
                if dark_text
                else colors.white
            )
            helper_color = (
                colors.HexColor("#66727C")
                if dark_text
                else colors.HexColor("#DDFBFF")
            )

            canvas.setFillColor(text_color)
            canvas.setFont("Helvetica-Bold", 8.5)
            canvas.drawCentredString(
                x + card_width / 2,
                y + card_height - 17,
                str(label)[:32],
            )

            value_size = 18
            if len(str(value)) > 13:
                value_size = 13
            elif len(str(value)) > 9:
                value_size = 15

            canvas.setFont("Helvetica-Bold", value_size)
            canvas.drawCentredString(
                x + card_width / 2,
                y + card_height / 2 - 2,
                str(value)[:24],
            )

            canvas.setFillColor(helper_color)
            canvas.setFont("Helvetica", 6.5)
            canvas.drawCentredString(
                x + card_width / 2,
                y + 10,
                str(helper)[:42],
            )


def _trend_chart(
    dataset: dict[str, Any],
    width: float,
    height: float,
) -> Drawing:
    labels, primary, secondary = _monthly_series(dataset)

    drawing = Drawing(width, height)
    drawing.add(
        Rect(
            0,
            0,
            width,
            height,
            rx=6,
            ry=6,
            fillColor=colors.white,
            strokeColor=colors.HexColor("#D9DEE2"),
        )
    )
    drawing.add(
        String(
            14,
            height - 19,
            "Activity trend",
            fontName="Helvetica-Bold",
            fontSize=10,
            fillColor=colors.HexColor("#27313A"),
        )
    )

    chart = HorizontalLineChart()
    chart.x = 45
    chart.y = 30
    chart.height = height - 65
    chart.width = width - 65
    chart.data = [primary, secondary]
    chart.categoryAxis.categoryNames = [
        label.replace(" 20", " '")
        for label in labels
    ]
    chart.categoryAxis.labels.fontSize = 6
    chart.categoryAxis.labels.angle = 20
    chart.categoryAxis.labels.dy = -7
    chart.valueAxis.labels.fontSize = 6
    chart.valueAxis.valueMin = 0
    chart.lines[0].strokeColor = colors.HexColor("#B46B6A")
    chart.lines[0].strokeWidth = 1.6
    chart.lines[1].strokeColor = colors.HexColor("#5B9B63")
    chart.lines[1].strokeWidth = 1.6
    chart.joinedLines = 1
    drawing.add(chart)

    return drawing


def _distribution_chart(
    dataset: dict[str, Any],
    width: float,
    height: float,
) -> Drawing:
    statuses = _status_counter(dataset)
    items = statuses.most_common(6)

    drawing = Drawing(width, height)
    drawing.add(
        Rect(
            0,
            0,
            width,
            height,
            rx=6,
            ry=6,
            fillColor=colors.white,
            strokeColor=colors.HexColor("#D9DEE2"),
        )
    )
    drawing.add(
        String(
            12,
            height - 19,
            "Status distribution",
            fontName="Helvetica-Bold",
            fontSize=10,
            fillColor=colors.HexColor("#27313A"),
        )
    )

    pie = Pie()
    pie.x = 25
    pie.y = 24
    pie.width = min(width * 0.42, 95)
    pie.height = min(height * 0.58, 95)
    pie.data = [
        count
        for _, count in items
    ] or [1]
    pie.labels = ["" for _ in pie.data]
    pie.slices.strokeWidth = 0.5

    palette = [
        colors.HexColor("#5B9B63"),
        colors.HexColor("#E75B5B"),
        colors.HexColor("#F4C04E"),
        colors.HexColor("#54B7C8"),
        colors.HexColor("#7B6CD1"),
        colors.HexColor("#9FA8B0"),
    ]
    for index in range(len(pie.data)):
        pie.slices[index].fillColor = palette[
            index % len(palette)
        ]
    drawing.add(pie)

    legend_x = width * 0.54
    legend_y = height - 42
    for index, (label, count) in enumerate(items):
        y = legend_y - index * 17
        drawing.add(
            Rect(
                legend_x,
                y - 2,
                8,
                8,
                fillColor=palette[index % len(palette)],
                strokeColor=None,
            )
        )
        drawing.add(
            String(
                legend_x + 13,
                y,
                f"{label}: {count}",
                fontName="Helvetica",
                fontSize=7,
                fillColor=colors.HexColor("#48535C"),
            )
        )

    return drawing


def _readiness_blocks(
    dataset: dict[str, Any],
    width: float,
    height: float,
) -> Drawing:
    statuses = _status_counter(dataset)
    total = sum(statuses.values()) or 1

    positive_words = {
        "active",
        "open",
        "paid",
        "complete",
        "completed",
        "won",
        "credit",
        "available",
        "ready",
    }
    negative_words = {
        "closed",
        "overdue",
        "past due",
        "inactive",
        "cancelled",
        "canceled",
        "lost",
        "charge",
        "not ready",
    }

    positive = sum(
        count
        for name, count in statuses.items()
        if name.casefold() in positive_words
    )
    negative = sum(
        count
        for name, count in statuses.items()
        if name.casefold() in negative_words
    )
    neutral = max(total - positive - negative, 0)

    drawing = Drawing(width, height)
    drawing.add(
        Rect(
            0,
            height / 2,
            width,
            height / 2,
            fillColor=colors.HexColor("#8FD24A"),
            strokeColor=colors.HexColor("#D9DEE2"),
        )
    )
    drawing.add(
        Rect(
            0,
            0,
            width,
            height / 2,
            fillColor=colors.HexColor("#F05A55"),
            strokeColor=colors.HexColor("#D9DEE2"),
        )
    )

    positive_label = (
        "Credits"
        if dataset.get("data_type") == "billing_ledger"
        else "Positive / Ready"
    )
    negative_label = (
        "Charges"
        if dataset.get("data_type") == "billing_ledger"
        else "Attention / Negative"
    )

    drawing.add(
        String(
            width / 2,
            height * 0.73,
            positive_label,
            textAnchor="middle",
            fontName="Helvetica-Bold",
            fontSize=9,
            fillColor=colors.HexColor("#16331B"),
        )
    )
    drawing.add(
        String(
            width / 2,
            height * 0.60,
            str(positive),
            textAnchor="middle",
            fontName="Helvetica-Bold",
            fontSize=17,
            fillColor=colors.HexColor("#16331B"),
        )
    )
    drawing.add(
        String(
            width / 2,
            height * 0.23,
            negative_label,
            textAnchor="middle",
            fontName="Helvetica-Bold",
            fontSize=9,
            fillColor=colors.white,
        )
    )
    drawing.add(
        String(
            width / 2,
            height * 0.10,
            str(negative),
            textAnchor="middle",
            fontName="Helvetica-Bold",
            fontSize=17,
            fillColor=colors.white,
        )
    )

    if neutral:
        drawing.add(
            String(
                width / 2,
                height / 2 - 3,
                f"Neutral: {neutral}",
                textAnchor="middle",
                fontName="Helvetica",
                fontSize=6,
                fillColor=colors.HexColor("#4A4A4A"),
            )
        )

    return drawing


def _dashboard_page(
    dataset: dict[str, Any],
    *,
    title: str,
    styles: dict[str, ParagraphStyle],
    page_width: float,
) -> list[Any]:
    elements: list[Any] = []

    elements.append(
        Paragraph(
            str(dataset.get("title") or title),
            styles["heading"],
        )
    )
    customer_name = (
        (dataset.get("data") or {}).get("customer_name")
        or ""
    )
    if customer_name:
        elements.append(
            Paragraph(
                f"Customer: {customer_name}",
                styles["body"],
            )
        )
    elements.append(Spacer(1, 7))

    elements.append(
        DashboardKPIGrid(
            _dashboard_metrics(dataset),
            page_width,
        )
    )
    elements.append(Spacer(1, 9))

    trend_width = page_width * 0.57
    distribution_width = page_width * 0.26
    readiness_width = page_width * 0.15
    chart_height = 2.15 * inch

    charts = Table(
        [
            [
                _trend_chart(
                    dataset,
                    trend_width,
                    chart_height,
                ),
                _distribution_chart(
                    dataset,
                    distribution_width,
                    chart_height,
                ),
                _readiness_blocks(
                    dataset,
                    readiness_width,
                    chart_height,
                ),
            ]
        ],
        colWidths=[
            trend_width,
            distribution_width,
            readiness_width,
        ],
    )
    charts.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    elements.append(charts)
    return elements



CCWR_STATUS_COLORS = {
    "EXPIRED": colors.HexColor("#86BFF2"),
    "ACTIVE": colors.HexColor("#172EAA"),
    "CANCELLED": colors.HexColor("#F5B99C"),
    "CANCELED": colors.HexColor("#F5B99C"),
    "OVERDUE": colors.HexColor("#B88BC4"),
}

CCWR_BUCKET_COLORS = {
    "Expired": colors.HexColor("#1689F4"),
    "365+ Days": colors.HexColor("#142CA8"),
    "181-365 Days": colors.HexColor("#EA6428"),
    "0-30 Days": colors.HexColor("#6E078C"),
    "91-180 Days": colors.HexColor("#D82E9B"),
    "61-90 Days": colors.HexColor("#7652C6"),
    "31-60 Days": colors.HexColor("#D9A800"),
}


def _ccwr_powerbi_bucket(record: dict[str, Any]) -> str:
    status = str(record.get("status") or "").upper()
    days = record.get("days_until_renewal")

    if status == "EXPIRED":
        return "Expired"
    if not isinstance(days, int):
        return "Unknown"
    if days < 0:
        return "Expired"
    if days <= 30:
        return "0-30 Days"
    if days <= 60:
        return "31-60 Days"
    if days <= 90:
        return "61-90 Days"
    if days <= 180:
        return "91-180 Days"
    if days <= 365:
        return "181-365 Days"
    return "365+ Days"


def _ccwr_dashboard_counts(
    records: list[dict[str, Any]],
) -> dict[str, int]:
    status_counts = Counter(
        str(record.get("status") or "UNKNOWN").upper()
        for record in records
    )
    overdue = sum(
        1
        for record in records
        if record.get("status") == "ACTIVE"
        and record.get("is_past_due") is True
    )
    return {
        "total": len(records),
        "active": status_counts.get("ACTIVE", 0),
        "overdue": overdue,
        "cancelled": (
            status_counts.get("CANCELLED", 0)
            + status_counts.get("CANCELED", 0)
        ),
        "expired": status_counts.get("EXPIRED", 0),
    }


class CcwrHeadlineCards(Flowable):
    def __init__(
        self,
        counts: dict[str, int],
        width: float,
    ):
        super().__init__()
        self.counts = counts
        self.width = width
        self.height = 1.05 * inch

    def wrap(self, avail_width, avail_height):
        return self.width, self.height

    def draw(self):
        canvas = self.canv
        gap = 10
        first_width = self.width * 0.16
        remaining_width = (
            self.width - first_width - gap * 4
        ) / 4

        cards = [
            (
                "",
                self.counts["total"],
                "Count of Subscription ID",
                first_width,
            ),
            (
                "ACTIVE",
                self.counts["active"],
                "Count of Subscription ID",
                remaining_width,
            ),
            (
                "OVERDUE",
                self.counts["overdue"],
                "Count of Subscription ID",
                remaining_width,
            ),
            (
                "CANCELLED",
                self.counts["cancelled"],
                "Count of Subscription ID",
                remaining_width,
            ),
            (
                "EXPIRED",
                self.counts["expired"],
                "Count of Subscription ID",
                remaining_width,
            ),
        ]

        x = 0
        for index, (heading, value, caption, width) in enumerate(cards):
            canvas.setFillColor(colors.white)
            canvas.setStrokeColor(colors.HexColor("#C8C5C2"))
            canvas.roundRect(
                x,
                0,
                width,
                self.height,
                3,
                fill=1,
                stroke=1,
            )

            if heading:
                canvas.setFillColor(colors.HexColor("#F0EFEE"))
                canvas.rect(
                    x,
                    self.height - 27,
                    width,
                    27,
                    fill=1,
                    stroke=0,
                )
                canvas.setStrokeColor(colors.HexColor("#C8C5C2"))
                canvas.line(
                    x,
                    self.height - 27,
                    x + width,
                    self.height - 27,
                )
                canvas.setFillColor(colors.HexColor("#313131"))
                canvas.setFont("Helvetica-Bold", 10)
                canvas.drawString(
                    x + 8,
                    self.height - 18,
                    heading,
                )
                number_y = self.height - 55
                caption_y = 13
            else:
                number_y = self.height - 43
                caption_y = 18

            canvas.setFillColor(colors.HexColor("#252525"))
            canvas.setFont("Helvetica", 21)
            canvas.drawString(
                x + 9,
                number_y,
                f"{value:,}",
            )
            canvas.setFillColor(colors.HexColor("#68635F"))
            canvas.setFont("Helvetica", 8.5)
            canvas.drawString(
                x + 9,
                caption_y,
                caption,
            )
            x += width + gap


def _ccwr_donut(
    *,
    title: str,
    values: list[int],
    labels: list[str],
    palette: list[Any],
    width: float,
    height: float,
) -> Drawing:
    drawing = Drawing(width, height)
    drawing.add(
        String(
            0,
            height - 14,
            title,
            fontName="Helvetica",
            fontSize=10,
            fillColor=colors.HexColor("#272727"),
        )
    )

    pie = Pie()
    pie.x = 30
    pie.y = 5
    pie.width = min(135, width * 0.48)
    pie.height = min(135, height - 30)
    pie.data = values or [1]
    pie.labels = ["" for _ in pie.data]
    pie.innerRadiusFraction = 0.58
    pie.slices.strokeWidth = 0.4

    for index, _ in enumerate(pie.data):
        pie.slices[index].fillColor = palette[index]
    drawing.add(pie)

    total = sum(values) or 1
    legend_x = width * 0.55
    legend_y = height - 45

    for index, (label, value) in enumerate(zip(labels, values)):
        y = legend_y - index * 16
        drawing.add(
            Rect(
                legend_x,
                y - 2,
                7,
                7,
                fillColor=palette[index],
                strokeColor=None,
            )
        )
        percentage = value / total * 100
        drawing.add(
            String(
                legend_x + 12,
                y,
                f"{label}: {value} ({percentage:.1f}%)",
                fontName="Helvetica",
                fontSize=6.8,
                fillColor=colors.HexColor("#55504C"),
            )
        )

    return drawing


def _ccwr_filter_chips(
    bucket_counts: Counter[str],
    width: float,
) -> Table:
    ordered = [
        "0-30 Days",
        "181-365 Days",
        "31-60 Days",
        "365+ Days",
        "61-90 Days",
        "91-180 Days",
        "Expired",
    ]

    cells = []
    for bucket in ordered:
        count = bucket_counts.get(bucket, 0)
        cells.append(
            Paragraph(
                f"<b>{bucket}</b><br/><font size='7'>{count:,} subscriptions</font>",
                ParagraphStyle(
                    f"Chip-{bucket}",
                    fontName="Helvetica",
                    fontSize=8.5,
                    leading=11,
                    textColor=colors.HexColor("#35322F"),
                    alignment=TA_LEFT,
                ),
            )
        )

    rows = [
        cells[0:3],
        cells[3:6],
        [cells[6], "", ""],
    ]
    table = Table(
        rows,
        colWidths=[width / 3] * 3,
        rowHeights=[0.55 * inch] * 3,
    )
    style = [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 9),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    for row in range(3):
        for col in range(3):
            if rows[row][col] != "":
                style.extend(
                    [
                        (
                            "BOX",
                            (col, row),
                            (col, row),
                            0.55,
                            colors.HexColor("#C8C5C2"),
                        ),
                        (
                            "BACKGROUND",
                            (col, row),
                            (col, row),
                            colors.white,
                        ),
                    ]
                )
    table.setStyle(TableStyle(style))
    return table


def _ccwr_detail_rows(
    records: list[dict[str, Any]],
) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "End Customer Name",
        "Subscription ID",
        "Year",
        "Quarter",
        "Month",
        "Day",
        "Auto Renew",
        "Status",
        "Renewal Bucket",
        "Risk",
    ]
    rows = []

    for record in records:
        renewal_text = (
            record.get("effective_renewal_date")
            or record.get("renewal_date")
            or record.get("end_date")
        )
        renewal_date = _date_value(renewal_text)

        rows.append(
            [
                record.get("end_customer_name") or "",
                record.get("subscription_id") or "",
                renewal_date.year if renewal_date else "",
                (
                    f"Qtr {(renewal_date.month - 1) // 3 + 1}"
                    if renewal_date
                    else ""
                ),
                renewal_date.strftime("%B") if renewal_date else "",
                renewal_date.day if renewal_date else "",
                "Yes" if record.get("has_auto_renewal") else "No",
                record.get("status") or "",
                _ccwr_powerbi_bucket(record),
                record.get("risk_level") or "",
            ]
        )

    return headers, rows


def _ccwr_pdf_dashboard(
    dataset: dict[str, Any],
    *,
    title: str,
    styles: dict[str, ParagraphStyle],
    usable_width: float,
) -> list[Any]:
    records = _dataset_records(dataset)
    counts = _ccwr_dashboard_counts(records)

    status_counter = Counter(
        "OVERDUE"
        if record.get("status") == "ACTIVE"
        and record.get("is_past_due") is True
        else str(record.get("status") or "UNKNOWN").upper()
        for record in records
    )
    status_order = [
        status
        for status in ["EXPIRED", "ACTIVE", "CANCELLED", "OVERDUE"]
        if status_counter.get(status)
    ]
    status_values = [status_counter[status] for status in status_order]
    status_palette = [
        CCWR_STATUS_COLORS.get(
            status,
            colors.HexColor("#999999"),
        )
        for status in status_order
    ]

    bucket_counter = Counter(
        _ccwr_powerbi_bucket(record)
        for record in records
    )
    bucket_order = [
        bucket
        for bucket in [
            "Expired",
            "365+ Days",
            "181-365 Days",
            "0-30 Days",
            "91-180 Days",
            "61-90 Days",
            "31-60 Days",
        ]
        if bucket_counter.get(bucket)
    ]
    bucket_values = [bucket_counter[bucket] for bucket in bucket_order]
    bucket_palette = [
        CCWR_BUCKET_COLORS.get(
            bucket,
            colors.HexColor("#999999"),
        )
        for bucket in bucket_order
    ]

    elements: list[Any] = [
        Paragraph(title, styles["title"]),
        Paragraph(
            "Cisco subscription renewal dashboard | "
            f"Generated {datetime.now(timezone.utc).strftime('%B %d, %Y at %I:%M %p UTC')}",
            styles["body"],
        ),
        Spacer(1, 7),
        CcwrHeadlineCards(counts, usable_width),
        Spacer(1, 12),
    ]

    chart_width = usable_width * 0.49
    chart_height = 2.05 * inch
    chart_table = Table(
        [
            [
                _ccwr_donut(
                    title=(
                        "Count of Subscription ID by Subscription Status"
                    ),
                    values=status_values,
                    labels=status_order,
                    palette=status_palette,
                    width=chart_width,
                    height=chart_height,
                ),
                _ccwr_donut(
                    title=(
                        "Count of Subscription ID by Renewal Bucket"
                    ),
                    values=bucket_values,
                    labels=bucket_order,
                    palette=bucket_palette,
                    width=chart_width,
                    height=chart_height,
                ),
            ]
        ],
        colWidths=[chart_width, chart_width],
    )
    chart_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    elements.append(chart_table)
    elements.append(Spacer(1, 8))

    elements.append(
        Paragraph(
            "Renewal Bucket",
            ParagraphStyle(
                "CcwrFilterTitle",
                parent=styles["heading"],
                fontSize=10,
                leading=12,
            ),
        )
    )
    elements.append(
        _ccwr_filter_chips(
            bucket_counter,
            usable_width,
        )
    )
    elements.append(PageBreak())

    headers, rows = _ccwr_detail_rows(records)
    table_data = [
        [
            Paragraph(
                str(header),
                ParagraphStyle(
                    f"CCWRHeader{index}",
                    fontName="Helvetica",
                    fontSize=6.5,
                    leading=7.5,
                    textColor=colors.HexColor("#292929"),
                ),
            )
            for index, header in enumerate(headers)
        ]
    ]

    for row in rows:
        table_data.append(
            [
                Paragraph(
                    str(value),
                    ParagraphStyle(
                        "CCWRCell",
                        fontName="Helvetica",
                        fontSize=6.1,
                        leading=7.2,
                        textColor=colors.HexColor("#37322F"),
                    ),
                )
                for value in row
            ]
        )

    widths = [
        usable_width * 0.31,
        usable_width * 0.10,
        usable_width * 0.055,
        usable_width * 0.065,
        usable_width * 0.075,
        usable_width * 0.045,
        usable_width * 0.07,
        usable_width * 0.08,
        usable_width * 0.11,
        usable_width * 0.09,
    ]
    detail_table = Table(
        table_data,
        colWidths=widths,
        repeatRows=1,
    )
    detail_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#F4F3F2"),
                ),
                (
                    "LINEBELOW",
                    (0, 0),
                    (-1, 0),
                    1,
                    colors.HexColor("#4A9BFF"),
                ),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [
                        colors.white,
                        colors.HexColor("#ECEBEA"),
                    ],
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.2,
                    colors.HexColor("#D5D3D1"),
                ),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    elements.append(
        Paragraph(
            "Renewal Subscription Detail",
            styles["heading"],
        )
    )
    elements.append(detail_table)
    return elements


def _ccwr_xlsx_dashboard(
    workbook: Workbook,
    *,
    datasets: list[dict[str, Any]],
    title: str,
) -> None:
    """
    Create an Excel dashboard that mirrors the Power BI-style renewal report.
    """
    dashboard = workbook.worksheets[0]
    dashboard.title = "Renewal Dashboard"

    records = [
        record
        for dataset in datasets
        if dataset.get("data_type") == "ccwr_renewals"
        for record in _dataset_records(dataset)
    ]
    counts = _ccwr_dashboard_counts(records)
    bucket_counts = Counter(
        _ccwr_powerbi_bucket(record)
        for record in records
    )
    status_counts = Counter(
        "OVERDUE"
        if record.get("status") == "ACTIVE"
        and record.get("is_past_due") is True
        else str(record.get("status") or "UNKNOWN").upper()
        for record in records
    )

    dashboard.sheet_view.showGridLines = False
    dashboard["A1"] = title
    dashboard["A1"].font = Font(
        size=20,
        bold=True,
        color="252525",
    )
    dashboard["A2"] = "Cisco subscription renewal dashboard"
    dashboard["A2"].font = Font(
        size=10,
        color="68635F",
    )

    cards = [
        ("TOTAL", counts["total"]),
        ("ACTIVE", counts["active"]),
        ("OVERDUE", counts["overdue"]),
        ("CANCELLED", counts["cancelled"]),
        ("EXPIRED", counts["expired"]),
    ]
    starts = ["A4", "D4", "G4", "J4", "M4"]
    for (label, value), start in zip(cards, starts):
        column = dashboard[start].column
        start_row = dashboard[start].row
        end_column = column + 2
        dashboard.merge_cells(
            start_row=start_row,
            start_column=column,
            end_row=start_row,
            end_column=end_column,
        )
        dashboard.merge_cells(
            start_row=start_row + 1,
            start_column=column,
            end_row=start_row + 2,
            end_column=end_column,
        )
        dashboard.cell(start_row, column, label)
        dashboard.cell(start_row + 1, column, value)
        dashboard.cell(start_row + 3, column, "Count of Subscription ID")
        for row in range(start_row, start_row + 4):
            for col in range(column, end_column + 1):
                cell = dashboard.cell(row, col)
                cell.fill = PatternFill(
                    "solid",
                    fgColor=(
                        "F0EFEE"
                        if row == start_row
                        else "FFFFFF"
                    ),
                )
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="left",
                )
                cell.border = __import__(
                    "openpyxl.styles",
                    fromlist=["Border", "Side"],
                ).Border(
                    left=__import__(
                        "openpyxl.styles",
                        fromlist=["Side"],
                    ).Side(style="thin", color="C8C5C2"),
                    right=__import__(
                        "openpyxl.styles",
                        fromlist=["Side"],
                    ).Side(style="thin", color="C8C5C2"),
                    top=__import__(
                        "openpyxl.styles",
                        fromlist=["Side"],
                    ).Side(style="thin", color="C8C5C2"),
                    bottom=__import__(
                        "openpyxl.styles",
                        fromlist=["Side"],
                    ).Side(style="thin", color="C8C5C2"),
                )
        dashboard.cell(start_row, column).font = Font(
            bold=True,
            size=11,
            color="313131",
        )
        dashboard.cell(start_row + 1, column).font = Font(
            size=22,
            color="252525",
        )
        dashboard.cell(start_row + 3, column).font = Font(
            size=9,
            color="68635F",
        )

    # Write chart support data away from the main visible area.
    status_start = 2
    dashboard["T1"] = "Subscription Status"
    dashboard["U1"] = "Count"
    for index, status in enumerate(
        ["EXPIRED", "ACTIVE", "CANCELLED", "OVERDUE"],
        start=status_start,
    ):
        dashboard[f"T{index}"] = status
        dashboard[f"U{index}"] = status_counts.get(status, 0)

    dashboard["W1"] = "Renewal Bucket"
    dashboard["X1"] = "Count"
    bucket_order = [
        "Expired",
        "365+ Days",
        "181-365 Days",
        "0-30 Days",
        "91-180 Days",
        "61-90 Days",
        "31-60 Days",
    ]
    for index, bucket in enumerate(bucket_order, start=2):
        dashboard[f"W{index}"] = bucket
        dashboard[f"X{index}"] = bucket_counts.get(bucket, 0)

    from openpyxl.chart import DoughnutChart, Reference
    from openpyxl.chart.label import DataLabelList

    status_chart = DoughnutChart()
    status_chart.title = (
        "Count of Subscription ID by Subscription Status"
    )
    status_chart.add_data(
        Reference(
            dashboard,
            min_col=21,
            min_row=1,
            max_row=5,
        ),
        titles_from_data=True,
    )
    status_chart.set_categories(
        Reference(
            dashboard,
            min_col=20,
            min_row=2,
            max_row=5,
        )
    )
    status_chart.height = 8
    status_chart.width = 12
    status_chart.holeSize = 55
    status_chart.dataLabels = DataLabelList()
    status_chart.dataLabels.showPercent = True
    status_chart.dataLabels.showLeaderLines = True
    dashboard.add_chart(status_chart, "A9")

    bucket_chart = DoughnutChart()
    bucket_chart.title = (
        "Count of Subscription ID by Renewal Bucket"
    )
    bucket_chart.add_data(
        Reference(
            dashboard,
            min_col=24,
            min_row=1,
            max_row=8,
        ),
        titles_from_data=True,
    )
    bucket_chart.set_categories(
        Reference(
            dashboard,
            min_col=23,
            min_row=2,
            max_row=8,
        )
    )
    bucket_chart.height = 8
    bucket_chart.width = 12
    bucket_chart.holeSize = 55
    bucket_chart.dataLabels = DataLabelList()
    bucket_chart.dataLabels.showPercent = True
    bucket_chart.dataLabels.showLeaderLines = True
    dashboard.add_chart(bucket_chart, "I9")

    # Renewal bucket "slicer-like" cards.
    dashboard["M9"] = "Renewal Bucket"
    dashboard["M9"].font = Font(
        bold=True,
        size=11,
        color="4A4541",
    )
    chip_cells = [
        ("M10", "0-30 Days"),
        ("P10", "181-365 Days"),
        ("S10", "31-60 Days"),
        ("M13", "365+ Days"),
        ("P13", "61-90 Days"),
        ("S13", "91-180 Days"),
        ("M16", "Expired"),
    ]
    for cell_ref, bucket in chip_cells:
        cell = dashboard[cell_ref]
        cell.value = (
            f"{bucket}\n{bucket_counts.get(bucket, 0)} subscriptions"
        )
        cell.alignment = Alignment(
            wrap_text=True,
            vertical="center",
            horizontal="left",
        )
        cell.font = Font(
            size=10,
            color="35322F",
        )
        cell.fill = PatternFill(
            "solid",
            fgColor="FFFFFF",
        )
        # Merge a 3-column x 2-row card.
        row = cell.row
        col = cell.column
        dashboard.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row + 1,
            end_column=col + 2,
        )

    headers, rows = _ccwr_detail_rows(records)
    detail = workbook.create_sheet("Renewal Detail")
    detail.append(headers)
    for row in rows:
        detail.append(row)

    header_fill = PatternFill("solid", fgColor="F0EFEE")
    for cell in detail[1]:
        cell.font = Font(
            bold=True,
            color="292929",
        )
        cell.fill = header_fill
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )

    widths = [42, 18, 9, 10, 13, 8, 12, 13, 18, 13]
    for index, width in enumerate(widths, start=1):
        detail.column_dimensions[
            get_column_letter(index)
        ].width = width

    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions

    for row_index in range(2, detail.max_row + 1):
        if row_index % 2 == 0:
            for col_index in range(1, detail.max_column + 1):
                detail.cell(row_index, col_index).fill = PatternFill(
                    "solid",
                    fgColor="ECEBEA",
                )

    dashboard.column_dimensions["T"].hidden = True
    dashboard.column_dimensions["U"].hidden = True
    dashboard.column_dimensions["W"].hidden = True
    dashboard.column_dimensions["X"].hidden = True

    for col in range(1, 22):
        dashboard.column_dimensions[
            get_column_letter(col)
        ].width = 11


def _create_pdf(
    path: Path,
    *,
    datasets: list[dict[str, Any]],
    title: str,
    template: str,
    include_summary: bool,
    include_raw_records: bool,
) -> None:
    styles = _pdf_styles()

    # Dashboard-style reports need landscape space for KPI tiles and charts.
    page_size = landscape(letter)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=page_size,
        rightMargin=0.35 * inch,
        leftMargin=0.35 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.4 * inch,
        title=title,
    )

    usable_width = (
        page_size[0]
        - doc.leftMargin
        - doc.rightMargin
    )

    ccwr_dashboard_only = (
        bool(datasets)
        and all(
            dataset.get("data_type") == "ccwr_renewals"
            for dataset in datasets
        )
        and template in {"executive", "customer_facing"}
    )

    if ccwr_dashboard_only:
        story: list[Any] = []
    else:
        story = [
            Paragraph(title, styles["title"]),
            Paragraph(
                f"Generated "
                f"{datetime.now(timezone.utc).strftime('%B %d, %Y at %I:%M %p UTC')}",
                styles["body"],
            ),
            Spacer(1, 7),
        ]

        if include_summary:
            story.append(
                Paragraph(
                    TEMPLATE_DESCRIPTIONS[template],
                    styles["body"],
                )
            )
            story.append(Spacer(1, 9))

    for dataset_index, dataset in enumerate(datasets):
        if dataset_index:
            story.append(PageBreak())

        if (
            dataset.get("data_type") == "ccwr_renewals"
            and template in {"executive", "customer_facing"}
        ):
            story.extend(
                _ccwr_pdf_dashboard(
                    dataset,
                    title=title,
                    styles=styles,
                    usable_width=usable_width,
                )
            )
            continue

        # Executive and customer-facing reports begin with a dashboard page.
        if template in {"executive", "customer_facing"}:
            story.extend(
                _dashboard_page(
                    dataset,
                    title=title,
                    styles=styles,
                    page_width=usable_width,
                )
            )

            records = _dataset_records(dataset)
            if records:
                story.append(PageBreak())
                story.append(
                    Paragraph(
                        "Supporting details",
                        styles["heading"],
                    )
                )
        else:
            story.append(
                Paragraph(
                    str(dataset.get("title") or "Dataset"),
                    styles["heading"],
                )
            )

        if template == "audit":
            metadata = [
                ["Dataset ID", dataset.get("dataset_id")],
                ["Conversation ID", dataset.get("conversation_id")],
                ["Data type", dataset.get("data_type")],
                ["Source", dataset.get("source")],
                ["Original query", dataset.get("query")],
                ["Intent", dataset.get("intent")],
                ["Created at", dataset.get("created_at")],
                ["Record count", dataset.get("record_count")],
            ]
            audit_table = Table(
                [
                    [
                        Paragraph(str(label), styles["small"]),
                        Paragraph(str(value or ""), styles["small"]),
                    ]
                    for label, value in metadata
                ],
                colWidths=[1.35 * inch, usable_width - 1.35 * inch],
            )
            audit_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (0, -1), LIGHT_GREEN),
                        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#BBD2C2")),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 4),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            story.append(audit_table)
            story.append(Spacer(1, 10))
        elif template == "detailed":
            story.append(
                Paragraph(
                    f"Type: {dataset.get('data_type')} | "
                    f"Source: {dataset.get('source')} | "
                    f"Records: {dataset.get('record_count', 0)}",
                    styles["body"],
                )
            )
            story.append(Spacer(1, 8))

        records = _dataset_records(dataset)
        if not records:
            story.append(
                Paragraph(
                    "No tabular records were available.",
                    styles["body"],
                )
            )
            continue

        limit = _record_limit(
            template,
            include_raw_records=include_raw_records,
        )
        selected_records = (
            records
            if limit is None
            else records[:limit]
        )
        rows = [
            _flatten_record(
                record,
                template=template,
            )
            for record in selected_records
        ]
        headers = list(
            dict.fromkeys(
                key
                for record in rows
                for key in record.keys()
            )
        )

        if template == "executive":
            headers = headers[:7]
        elif template == "customer_facing":
            headers = headers[:8]
        elif template == "detailed":
            headers = headers[:12]
        else:
            headers = headers[:14]

        if not headers:
            story.append(
                Paragraph(
                    "No report-safe fields were available.",
                    styles["body"],
                )
            )
            continue

        table_data = [
            [
                Paragraph(
                    _friendly_label(str(header)),
                    styles["small"],
                )
                for header in headers
            ]
        ]
        for record in rows:
            table_data.append(
                [
                    Paragraph(
                        str(record.get(header, ""))[:550],
                        styles["small"],
                    )
                    for header in headers
                ]
            )

        col_width = usable_width / max(
            len(headers),
            1,
        )
        table = Table(
            table_data,
            colWidths=[col_width] * len(headers),
            repeatRows=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), DARK_GREEN),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C9D7CE")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F8F4")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)

        if template == "audit":
            story.append(Spacer(1, 12))
            story.append(
                Paragraph(
                    "Raw Dataset Evidence",
                    styles["heading"],
                )
            )
            raw_text = json.dumps(
                dataset.get("data") or {},
                ensure_ascii=False,
                default=str,
                indent=2,
            )
            story.append(
                Paragraph(
                    raw_text[:20_000]
                    .replace("&", "&amp;")
                    .replace("<", "&lt;")
                    .replace(">", "&gt;")
                    .replace("\n", "<br/>"),
                    styles["small"],
                )
            )

    doc.build(story)
